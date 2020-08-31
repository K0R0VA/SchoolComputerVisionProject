# coding=Windows-1251
import os
from random import randint
import openpyxl
import cv2
import numpy as np
from kivy import factory
from kivy.app import App
from kivy.core.image import Texture
from kivy.properties import ObjectProperty
from kivy.clock import Clock
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.popup import Popup
from kivy.uix.screenmanager import Screen, ScreenManager
import SliderTest
import Functions
from datetime import date
import Seconds
from stopwatch import Stopwatch

stop = Stopwatch()
stop.stop()
videoname = "NarfuSight_" + str(date.today()) + ".avi"
FileName = str(randint(1, 999)) + '_coordinates'


class MainScreen(Screen):
    pass


class OneCamScreen(Screen):
    wrapper = ObjectProperty(None)


class TwoCamScreen(Screen):
    wrapper2 = ObjectProperty(None)


class AnalyseScreen(Screen):
    def show3d(self):
        Functions.Show3D(FileName)
    def show3d2(self):
        Functions.Show3D2(FileName)
    def ShowSpeedChart(self):
        #r_high1 = App.get_running_app().root.onecam.wrapper.hue.value
        #g_high1 = App.get_running_app().root.onecam.wrapper.saturation.value
        #b_high1 = App.get_running_app().root.onecam.wrapper.value.value
        #r_high2 = App.get_running_app().root.onecam.wrapper.hue2.value
        #g_high2 = App.get_running_app().root.onecam.wrapper.saturation2.value
        #b_high2 = App.get_running_app().root.onecam.wrapper.value2.value
        #obj_clr1 = r_high1,g_high1,b_high1
        #obj_clr2 = r_high2, g_high2, b_high2
        Functions.ShowSpeedChart(FileName)

    def StartExcel(self):
        File= os.getcwd()  + FileName+ '.xlsx'
        os.system("start " + File)

    def StartVideo(self):
        Functions.Xt(FileName)
    def xt_yt(self):
        Functions.Xt_Yt(FileName)
class Popup_content(BoxLayout):
    obj1h = ObjectProperty(None)
    obj2h = ObjectProperty(None)
    obj1w = ObjectProperty(None)
    obj2w = ObjectProperty(None)

    def __init__(self, **kwargs):
        super(Popup_content, self).__init__(**kwargs)
        self.select_is_active = False

    save_obj_param_btn = ObjectProperty(None)


class CalibrateCam1Screen(Screen):
    popup = ObjectProperty(None)
    qr_h = ObjectProperty(None)
    qr_w = ObjectProperty(None)
    qr_dist = ObjectProperty(None)
    popupWindow = ObjectProperty(None)
    heightFirst = ObjectProperty(None)
    widthFirst = ObjectProperty(None)
    widthSecond = ObjectProperty(None)
    heightSecond = ObjectProperty(None)

    def __init__(self, **kwargs):
        super(CalibrateCam1Screen, self).__init__(**kwargs)

    cam = ObjectProperty(None)




class CalibrateCam2Screen(Screen):
    pass


class ScreenManager(ScreenManager):
    scrncond = ObjectProperty(None)
    main = ObjectProperty(None)
    onecam = ObjectProperty(None)
    twocam = ObjectProperty(None)
    wrap = ObjectProperty(None)
    analyse = ObjectProperty(None)
    cal1 = ObjectProperty(None)
    cal2 = ObjectProperty(None)


class Wrapper(BoxLayout):
    colorpick = ObjectProperty(None)
    colorpick2 = ObjectProperty(None)
    seconds = ObjectProperty(None)
    cam1 = ObjectProperty(None)
    recbtn = ObjectProperty(None)
    save_btn = ObjectProperty(None)
    r_high1 = ObjectProperty(None)
    g_high1 = ObjectProperty(None)
    b_high1 = ObjectProperty(None)
    r_high2 = ObjectProperty(None)
    g_high2 = ObjectProperty(None)
    b_high2 = ObjectProperty(None)
    def __init__(self, **kwargs):
        super(Wrapper, self).__init__(**kwargs)
        self.rec_is_active = False
        self.saved = False

    def record(self):
        self.rec_is_active = not self.rec_is_active
        if self.rec_is_active:
            self.recbtn.text = 'Пауза'
            stop.start()
        else:
            self.recbtn.text = 'Запись'
            stop.stop()

    def savebtn(self):
        self.saved = not self.saved
        if self.saved:
            self.save_btn.text = 'Сохранено'
        else:
            self.save_btn.text = 'Сохранить'

    def SaveExel(self):
        pass
        # print(CamMainApp.rec_path)
        # Functions.MakeNewExcelFile(recs, cv2.os.getcwd(), 'coords')

    def ResetWatch(self):
        stop.reset()
        self.seconds.text = str(stop)
        self.recbtn.text = 'Запись'
        self.rec_is_active = False
        with open(videoname, "w"):
            pass

    def ColorPicklistener(self):
        if self.hue.value[1] != 0:
            r_high1 = round((100 / (255 / self.hue.value[1]) / 100), 2)
        else:
            r_high1 = 0
        if self.saturation.value[1] != 0:
            g_high1 = round((100 / (255 / self.saturation.value[1]) / 100), 2)
        else:
            g_high1 = 0
        if self.value.value[1] != 0:
            b_high1 = round((100 / (255 / self.value.value[1]) / 100), 2)
        else:
            b_high1 = 0
        if self.hue.value[0] != 0:
            r_low1 = round((100 / (255 / self.hue.value[0]) / 100), 2)
        else:
            r_low1 = 0
        if self.saturation.value[0] != 0:
            g_low1 = round((100 / (255 / self.saturation.value[0]) / 100), 2)
        else:
            g_low1 = 0
        if self.value.value[0] != 0:
            b_low1 = round((100 / (255 / self.value.value[0]) / 100), 2)
        else:
            b_low1 = 0
        if self.hue2.value[1] != 0:
            r_high2 = round((100 / (255 / self.hue2.value[1]) / 100), 2)
        else:
            r_high2 = 0
        if self.saturation2.value[1] != 0:
            g_high2 = round((100 / (255 / self.saturation2.value[1]) / 100), 2)
        else:
            g_high2 = 0
        if self.value2.value[1] != 0:
            b_high2 = round((100 / (255 / self.value2.value[1]) / 100), 2)
        else:
            b_high2 = 0
        if self.hue2.value[0] != 0:
            r_low2 = round((100 / (255 / self.hue2.value[0]) / 100), 2)
        else:
            r_low2 = 0
        if self.saturation2.value[0] != 0:
            g_low2 = round((100 / (255 / self.saturation2.value[0]) / 100), 2)
        else:
            g_low2 = 0
        if self.value2.value[0] != 0:
            b_low2 = round((100 / (255 / self.value2.value[0]) / 100), 2)
        else:
            b_low2 = 0
        self.colorpick.color = r_high1, g_high1, b_high1, 1
        self.colorpick2.color = r_low1, g_low1, b_low1, 1
        self.colorpick3.color = r_high2, g_high2, b_high2, 1
        self.colorpick4.color = r_low2, g_low2, b_low2, 1


class Wrapper2(BoxLayout):
    def __init__(self, **kwargs):
        super(Wrapper2, self).__init__(**kwargs)


class CamMainApp(App):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.capture = cv2.VideoCapture(1)
        self.capture2 = cv2.VideoCapture(0)
        self.PathOfFirstobject = []
        self.PathOfSecondobject = []
        self.delta = 4
        self.distance = 1
        # Параметры видеозаписи
        self.fourcc = cv2.VideoWriter_fourcc(*'XVID')
        self.w = int(self.capture.get(cv2.CAP_PROP_FRAME_WIDTH))
        self.h = int(self.capture.get(cv2.CAP_PROP_FRAME_HEIGHT))
        self.video_writer = cv2.VideoWriter(videoname, self.fourcc, 12, (self.w, self.h))
        self.sec = 0
        self.coord_count = 0
        self.rec_pathfirstobject = []
        self.rec_pathsecondobject = []
        self.timer = []
        self.workbook = openpyxl.Workbook()
        self.hypotenuseFirst = 1
        self.hypotenuseSecond = 1

    def build(self):
        Clock.schedule_interval(self.update, 1.0 / 33.0)
        return ScreenManager()

    def MakeNewExcelFile(self):
        ws = self.workbook.active
        ws.cell(1, 1).value = 'X axis (мм)'
        ws.cell(1, 2).value = 'Y axis (мм)'
        ws.cell(1, 3).value = 'Z axis (мм)'
        ws.cell(1, 4).value = 'Time (сек)'
        k = (2 * self.rec_pathfirstobject[0][2]) / self.hypotenuseFirst
        for i in range(1, len(self.rec_pathfirstobject)):
            if self.rec_pathfirstobject[i][2] != 0 and self.hypotenuseFirst != 0:
                r = self.rec_pathfirstobject[i][2] / self.rec_pathfirstobject[0][2]
                ws.cell(i + 1, 1).value = int((self.rec_pathfirstobject[i][0]) / (self.delta / k))
                ws.cell(i + 1, 2).value = int((self.rec_pathfirstobject[i][1]) / (self.delta / k))
                ws.cell(i + 1, 3).value = int(int(self.distance) / r)
                ws.cell(i + 1, 4).value = self.timer[i]
        ws.cell(1, 6).value = 'X axis (мм)'
        ws.cell(1, 7).value = 'Y axis (мм)'
        ws.cell(1, 8).value = 'Z axis (мм)'
        ws.cell(1, 9).value = 'Time (с)'
        k = self.delta / (2 * self.rec_pathsecondobject[i][2] / self.hypotenuseSecond)
        for i in range(1, len(self.rec_pathsecondobject)):
            if self.rec_pathsecondobject[i][2] != 0 and self.hypotenuseSecond != 0:
                r = self.rec_pathfirstobject[i][2] / self.rec_pathfirstobject[0][2]
                ws.cell(i + 1, 6).value = int(self.rec_pathsecondobject[i][0] / (self.delta / k))
                ws.cell(i + 1, 7).value = int(self.rec_pathsecondobject[i][1] / (self.delta / k))
                #ws.cell(i + 1, 8).value = int(int(self.distance) / r)
                ws.cell(i + 1, 9).value = self.timer[i]

    def update(self, dt):
        # Код для калибровки 1 камеры
        if App.get_running_app().root.scrncond.text == 'cal1':
            qr_w = int(App.get_running_app().root.cal1.qr_w.text)
            qr_h = int(App.get_running_app().root.cal1.qr_h.text)
            _, frame = self.capture.read()
            self.delta = Functions.detectMark(frame, qr_w, qr_h)
            texture = Texture.create(size=(frame.shape[1], frame.shape[0]), colorfmt='bgr')
            result = cv2.flip(frame, 0)
            texture.blit_buffer(result.tostring(), colorfmt='bgr', bufferfmt='ubyte')
            App.get_running_app().root.cal1.cam.texture = texture
        # Код для одной камеры, экран выделения
        if App.get_running_app().root.scrncond.text == 'onecam':
            _, frame = self.capture.read()
            result = cv2.flip(frame, 0)
            texture = Texture.create(size=(frame.shape[1], frame.shape[0]), colorfmt='bgr')
            H_Firstobj = App.get_running_app().root.onecam.wrapper.hue.value
            H_Secondobj = App.get_running_app().root.onecam.wrapper.hue2.value
            saturationFirst = App.get_running_app().root.onecam.wrapper.saturation.value
            saturationSecond = App.get_running_app().root.onecam.wrapper.saturation2.value
            valueFirst = App.get_running_app().root.onecam.wrapper.value.value
            valueSecond = App.get_running_app().root.onecam.wrapper.value2.value
            lowFirst = np.array([H_Firstobj[0], saturationFirst[0], valueFirst[0]])
            highFirst = np.array([H_Firstobj[1], saturationFirst[1], valueFirst[1]])
            lowSecond = np.array([H_Secondobj[0], saturationSecond[0], valueSecond[0]])
            highSecond = np.array([H_Secondobj[1], saturationSecond[1], valueSecond[1]])
            blur = int(App.get_running_app().root.onecam.wrapper.blur.value)
            masksecondobject = Functions.getMask(frame, lowSecond, highSecond, blur)
            maskfirstobject = Functions.getMask(frame, lowFirst, highFirst, blur)
            if saturationFirst == [0, 0] and H_Firstobj == [0, 0] and valueFirst == [0, 0]:
                cv2.destroyWindow('Mask1')
            else:
                Functions.getPoint(maskfirstobject, result, self.PathOfFirstobject, (0, 255, 0))
                cv2.imshow('Mask1', maskfirstobject)
            if saturationSecond == [0, 0] and H_Secondobj == [0, 0] and valueSecond == [0, 0]:

                cv2.destroyWindow('Mask2')
            else:
                Functions.getPoint(masksecondobject, result, self.PathOfSecondobject, (0, 0, 255))
                cv2.imshow('Mask2', masksecondobject)
            texture.blit_buffer(result.tostring(), colorfmt='bgr', bufferfmt='ubyte')
            App.get_running_app().root.onecam.wrapper.cam1.texture = texture

            # Начать запить видео
            if App.get_running_app().root.onecam.wrapper.recbtn.text == 'Пауза':
                if saturationSecond == [0, 0] and H_Secondobj == [0, 0] and valueSecond == [0, 0]:
                    self.rec_pathsecondobject.append([-999, -999, -999])
                else:
                    Functions.getPoint(masksecondobject, result, self.rec_pathsecondobject, (0, 0, 255))
                if saturationFirst == [0, 0] and H_Firstobj == [0, 0] and valueFirst == [0, 0]:
                    self.rec_pathfirstobject.append([-999, -999, -999])
                else:
                    Functions.getPoint(maskfirstobject, result, self.rec_pathfirstobject, (0, 255, 0))
                print(len(self.rec_pathsecondobject),len(self.rec_pathfirstobject))


                self.video_writer.write(result)
                record_is_running = True
                self.timer.append(str(stop))
                self.coord_count += self.coord_count
                App.get_running_app().root.onecam.wrapper.seconds.text = str(stop)
            else:
                record_is_running = False
        if App.get_running_app().root.onecam.wrapper.save_btn.text == 'Сохранено' and App.get_running_app().root.scrncond.text == 'analyse':
            FilePath = cv2.os.getcwd()
            self.distance = App.get_running_app().root.cal1.qr_dist.text
            Param = FilePath + FileName + '.xlsx'
            heightfirst = float(App.get_running_app().root.cal1.heightFirst.text)
            widthfirst = float(App.get_running_app().root.cal1.widthFirst.text)
            self.hypotenuseFirst = float(Functions.hypotenuse(heightfirst, widthfirst))
            heightsecond = float(App.get_running_app().root.cal1.heightSecond.text)
            widthsecond = float(App.get_running_app().root.cal1.widthSecond.text)
            self.hypotenuseSecond = float(Functions.hypotenuse(heightsecond, widthsecond))
            self.MakeNewExcelFile()
            self.workbook.save(Param)
            App.get_running_app().root.onecam.wrapper.save_btn.text = 'Сохранить'
